import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from datetime import datetime

# Configure Chrome options
def configure_chrome_options():
    chrome_options = Options()
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--ignore-ssl-errors")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-insecure-localhost")
    chrome_options.add_argument("--headless")
    return chrome_options

# Set up WebDriver
def setup_driver():
    service = Service("driver/chromedriver.exe")  
    chrome_options = configure_chrome_options()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()
    return driver

def scroll_results(driver):
    element_scroll = driver.find_element(By.XPATH, "//*[@id='QA0Szd']/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]")
    driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", element_scroll)
    result = driver.find_element(By.CLASS_NAME, "lMbq3e")
    driver.execute_script("arguments[0].scrollIntoView(true);", result)
    time.sleep(3)
                                         
def scroll_until_class_found(driver, target_class_name):
    element_to_scroll = driver.find_element(By.XPATH, "//*[@id='QA0Szd']/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]")

    while True:
        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", element_to_scroll)
        time.sleep(5)

        try:
            text = driver.find_element(By.CLASS_NAME, target_class_name).text
            print(f"Found target class: {text}")
            break
        except:
            pass
    driver.execute_script("window.scrollTo(0, 0)")

def extract_phone_numbers(elements):
    phone_numbers = []
    for element in elements:
        text = element.text.strip()
        if text:
            time.sleep(1)  # Pause for 1 second (if needed)
            phone_numbers.append(text)
    return phone_numbers

def find_and_format_phone_number(phone_numbers, pattern=r'\d{4}-\d{4}-\d{4}'):
    for item in phone_numbers:
        match = re.search(pattern, item)
        rematch = re.search(r'\d{4}-\d{4}-\d{3}', item)
        if match:
            return match.group().replace('-', '')
        elif rematch:
            return rematch.group().replace('-', '')
    return None

def process_search_results(driver, processed_names, data_list):
    results = driver.find_elements(By.CLASS_NAME, "Nv2PK")
    print(f"Found {len(results)} results.")
    
    for result in results:
        try:
            name = result.find_element(By.CLASS_NAME, "qBF1Pd").text
            if name in processed_names:
                print("Duplicate found, skipping...")
            processed_names.add(name)
            print(f"Name: {name}")

            try:
                result.click()
                scroll_results(driver)
                time.sleep(2)

                address = driver.find_element(By.CLASS_NAME, "fdkmkc").text
                find_phone_classes = driver.find_elements(By.CLASS_NAME, "Io6YTe")
                phone_numbers = extract_phone_numbers(find_phone_classes)
                phone = find_and_format_phone_number(phone_numbers)
                phone = phone if phone else 'Not available'
                address = address if address else 'Not available'
                
                print(f"Address: {address if address else 'Not available'}")
                print(f"Phone: {phone if phone else 'Not available'}")
                print("-" * 50)  

                data_list.append([name, address, phone])

            except:
                print("Scrolling to bring the result to the top...")
                driver.execute_script("arguments[0].scrollIntoView(true);", result)
                time.sleep(1)
                
                result.click()
                scroll_results(driver)
                time.sleep(1)
                
                address = driver.find_element(By.CLASS_NAME, "fdkmkc").text
                find_phone_classes = driver.find_elements(By.CLASS_NAME, "Io6YTe")
                phone_numbers = extract_phone_numbers(find_phone_classes)
                phone = find_and_format_phone_number(phone_numbers)
                phone = phone if phone else 'Not available'
                address = address if address else 'Not available'
                
                print(f"Address: {address if address else 'Not available'}")
                print(f"Phone: {phone if phone else 'Not available'}")
                print("-" * 50)   

                data_list.append([name, address, phone])

            time.sleep(2)

        except Exception as e:
            print(f"Error processing result: {e}")
            continue

def save_to_excel(data_list,file_name):
    # Create a pandas DataFrame
    df = pd.DataFrame(data_list, columns=['Name', 'Address', 'Phone'])
    df.index.name = 'No'
    df.index += 1
    # Write DataFrame to Excel
    df.to_excel(file_name, index=True)
    
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # Define yellow fill style
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    for row in sheet.iter_rows():
        for cell in row:
            # Apply yellow fill to non-empty cells in the first row
            if cell.row == 1 and cell.value:
                cell.fill = yellow_fill
            # Apply border to all cells
            cell.border = thin_border

    # Save the styled workbook
    workbook.save(file_name)

def main():
    driver = setup_driver()

    try:
        driver.get("https://www.google.com/maps")
        wait = WebDriverWait(driver, 20)

        search_keyword = "laundry di padangsambian"
        search_box = wait.until(EC.presence_of_element_located((By.ID, "searchboxinput")))
        search_box.send_keys(search_keyword)
        search_box.send_keys(Keys.ENTER)

        time.sleep(5)
        scroll_until_class_found(driver, "PbZDve")
        
        
        now = datetime.now()
        formatted_date_time = now.strftime("%Y%m%d_%H%M")
        processed_names = set()
        data_list = []
        process_search_results(driver, processed_names, data_list)
        file_name = f"{search_keyword.replace(' ','_')}_{formatted_date_time}.xlsx"


        # Save the collected data to Excel
        save_to_excel(data_list, str(file_name))

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
